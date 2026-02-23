"""
Exportação para Excel no formato de planilha de carga Qstione
Adaptado para funcionar com dados provenientes dos importadores SQL Server.
"""

import os
from datetime import datetime
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExportadorExcel:
    """
    Exporta dados das tabelas de importação para planilhas Excel no formato de carga.
    """

    def __init__(self, caminho_saida: str = './exportacoes'):
        self.caminho_saida = caminho_saida
        os.makedirs(caminho_saida, exist_ok=True)

    def criar_planilha_carga(
        self,
        nome_tabela: str,
        dados: List[Dict[str, Any]],
        config_tabela: Dict[str, Any]
    ) -> str:
        """
        Cria uma planilha Excel para uma tabela de importação.

        Args:
            nome_tabela: Nome da tabela (ex: 'imp_001_cursos')
            dados: Lista de dicionários com os dados a exportar (pode ser vazia)
            config_tabela: Configuração da tabela (contém 'campos', 'tipo_carga', etc.)

        Returns:
            Caminho completo do arquivo gerado.
        """
        # Garantir que dados seja uma lista (mesmo que vazia)
        if not isinstance(dados, list):
            dados = []

        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo = f"{nome_tabela}_{timestamp}.xlsx"
        caminho_arquivo = os.path.join(self.caminho_saida, nome_arquivo)

        # Cria workbook
        wb = openpyxl.Workbook()

        # ---------- Aba "Dados a Importar" ----------
        ws_dados = wb.active
        ws_dados.title = "Dados a Importar"

        # Estilos
        fonte_titulo = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
        alinhamento_direita = Alignment(horizontal='right', vertical='center')
        preenchimento_titulo = PatternFill(fill_type='solid', start_color='366092', end_color='366092')
        preenchimento_cabecalho = PatternFill(fill_type='solid', start_color='C5D9F1', end_color='C5D9F1')

        # Cabeçalho da planilha (metadados)
        cabecalho = [
            ["Tipo de Carga:", config_tabela.get('tipo_carga', 'Incremental')],
            ["Escopo da Carga:", config_tabela.get('escopo_carga', 'Instituicao')],
            ["Limite do Escopo:", ""],
            ["Conteúdo da Carga:", config_tabela.get('nome_planilha', nome_tabela)],
            [""],  # linha separadora
        ]

        for linha_idx, linha in enumerate(cabecalho, start=1):
            for col_idx, valor in enumerate(linha, start=1):
                celula = ws_dados.cell(row=linha_idx, column=col_idx, value=valor)
                if col_idx == 1 and linha_idx <= 4:  # títulos (coluna A)
                    celula.font = fonte_titulo
                    celula.alignment = alinhamento_direita
                    celula.fill = preenchimento_titulo
                elif col_idx == 2 and linha_idx <= 4:  # valores (coluna B)
                    celula.font = Font(name='Calibri', size=11, bold=False)
                    celula.alignment = alinhamento_esquerda

        # Nomes das colunas (obtidos da configuração)
        colunas = [campo['nome_qstione'] for campo in config_tabela['campos']]
        linha_cabecalho = 6  # após as 5 linhas de cabeçalho + 1 linha em branco

        for col_idx, coluna in enumerate(colunas, start=1):
            celula = ws_dados.cell(row=linha_cabecalho, column=col_idx, value=coluna)
            celula.font = Font(name='Calibri', size=11, bold=True)
            celula.fill = preenchimento_cabecalho
            celula.alignment = Alignment(horizontal='center', vertical='center')

        # Dados
        for linha_idx, registro in enumerate(dados, start=linha_cabecalho + 1):
            for col_idx, coluna in enumerate(colunas, start=1):
                valor = registro.get(coluna, '')  # se a chave não existir, string vazia
                ws_dados.cell(row=linha_idx, column=col_idx, value=valor)

        # Ajuste automático da largura das colunas (baseado no conteúdo do cabeçalho)
        for col_idx, coluna in enumerate(colunas, start=1):
            col_letter = get_column_letter(col_idx)
            # Largura fixa (pode ser ajustada conforme necessidade)
            ws_dados.column_dimensions[col_letter].width = 25

        # ---------- Aba "Configurações" ----------
        ws_config = wb.create_sheet(title="Configurações")
        linhas_config = [
            ["ATENÇÃO: Os dados desta planilha de Configurações não devem ser alterados, pois possuem apenas caráter informativo e de configurações para a planilha de Dados a Importar."],
            [""],
            ["Tipos de Carga"],
            ["Incremental", "Os dados constantes na planilha serão adicionados ou atualizarão os já existentes no sistema."],
            ["Completa", "Ocorre o comportamento previsto na Incremental e, adicionalmente, os dados que estejam no sistema, mas não na planilha, serão inativados."],
            [""],
            ["Escopos de Carga"],
            ["Instituicao", "Engloba todos os usuários da instituição de ensino."],
            ["Unidade Organizacional", "Engloba todos os cursos da unidade organizacional."],
            ["Curso", "Engloba todas as vinculações referentes ao curso com código preenchido no campo Limite do Escopo."],
            ["Disciplina", "Engloba todas as vinculações referentes à disciplina com código preenchido no campo Limite do Escopo."],
            ["Oferta de Disciplina", "Engloba todas as vinculações referentes à oferta de disciplina com código preenchido no campo Limite do Escopo."],
        ]

        for linha_idx, linha in enumerate(linhas_config, start=1):
            for col_idx, valor in enumerate(linha, start=1):
                ws_config.cell(row=linha_idx, column=col_idx, value=valor)

        ws_config.column_dimensions['A'].width = 50
        ws_config.column_dimensions['B'].width = 70

        # Salva o arquivo
        try:
            wb.save(caminho_arquivo)
        except Exception as e:
            print(f"  ✗ Erro ao salvar planilha {nome_arquivo}: {e}")
            return None

        # Mensagem de resultado
        print(f"✓ Planilha gerada: {caminho_arquivo}")
        if dados:
            print(f"  Total de registros: {len(dados)}")
        else:
            print(f"  (Planilha vazia – apenas cabeçalho)")

        return caminho_arquivo

    def exportar_todas_tabelas(
        self,
        dados_por_tabela: Dict[str, List[Dict[str, Any]]],
        config_tabelas: Dict[str, Any]
    ) -> List[str]:
        """
        Exporta múltiplas tabelas para arquivos Excel separados.

        Args:
            dados_por_tabela: Dicionário com nome da tabela -> lista de registros
            config_tabelas: Configuração de todas as tabelas (TABELAS_CONFIG)

        Returns:
            Lista com os caminhos dos arquivos gerados.
        """
        arquivos_gerados = []

        for nome_tabela, dados in dados_por_tabela.items():
            if nome_tabela not in config_tabelas:
                print(f"⚠️  Tabela '{nome_tabela}' não encontrada na configuração. Ignorada.")
                continue

            config = config_tabelas[nome_tabela]
            arquivo = self.criar_planilha_carga(nome_tabela, dados, config)
            if arquivo:
                arquivos_gerados.append(arquivo)

        return arquivos_gerados